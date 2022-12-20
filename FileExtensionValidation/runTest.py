# -*- coding: utf-8 -*-
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
# import openpyxl module
import openpyxl
import logging

def testcaseRunner(driver, filename, expected_string):
        # driver=self.driver
        driver.get("https://sandbox.moodledemo.net/")
        driver.find_element(By.XPATH, "/html/body/div[2]/nav/div[2]/div[5]/div/span/a").click()
        driver.find_element("id","username").click()
        driver.find_element("id","username").clear()
        driver.find_element("id","username").send_keys('student')

        driver.find_element("id","password").click()
        driver.find_element("id","password").clear()

        driver.find_element("id","password").send_keys('sandbox')
        driver.find_element("id","loginbtn").click()

        # driver.get("https://sandbox.moodledemo.net/user/edit.php?id=4&returnto=profile")
        driver.find_element("link text","My first course").click()

        xpath = '/html/body/div[3]/div[5]/div/div[1]/div[1]/button'
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,xpath)))
        except:
            pass

        try:
            action = ActionChains(driver)
            action.click(on_element=driver.find_element(By.XPATH,xpath))
            action.perform()
            time.sleep(5)
        except:
            pass
        try:
            driver.find_element("link text","test_file_extension_validation").click()
        except:
            raise Exception("Sorry, you have to prepare enviromnent first!!!")

        driver.current_url

        driver.get("{}&action=editsubmission".format(driver.current_url))
        # dragdrop = driver.find_element_by_xpath('paste-the-full-xpath-here')
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,".fm-empty-container")))
            # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"/html/body/div[3]/div[4]/div/div[2]/div/section/div[2]/div/form/fieldset/div[2]/div/div[2]/fieldset/div[1]/div[4]/div[1]/div[2]")))
        dragdrop = driver.find_element(By.CSS_SELECTOR,".dndupload-target")

        cssSelector=".dndupload-message > div:nth-child(2)"
        # className='dndupload-arrow d-flex'
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR,cssSelector)))

        # ActionChains(driver).move_to_element(driver.find_element(By.CSS_SELECTOR,cssSelector)).perform()

        element= driver.find_element(By.CSS_SELECTOR,cssSelector).click()
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable(("link text","Upload a file")))
        driver.find_element("link text","Upload a file").click()
        # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.NAME,'repo_upload_file')))
        # driver.find_elements(By.NAME,'repo_upload_file')[0].click() --> not working
        # driver.find_elements(By.NAME,'repo_upload_file')[0].send_keys("img.png") 

        # driver.find_element("link text","Choose File").click()  #not working

        # driver.execute_script("document.getElementsByName('repo_upload_file')[0].click()") # working
        path = os.path.join(os.getcwd(),"test_data", filename)
        # pyautogui.write(path, interval=0.25) 
        # # pyautogui.press('enter')
        # pyautogui.press('return')


        driver.find_element(By.NAME,'repo_upload_file' ).send_keys(path)
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]")))

        driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]").click()

        #Assert: 
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+expected_string+"')]")))
        driver.quit()


class AppDynamicsJob(unittest.TestCase):
    def setUp(self):

        options = FirefoxOptions()
        options.set_preference('intl.accept_languages', 'en-US, en')
        self.driver = webdriver.Firefox(options=options)


        # options = ChromeOptions()
        # options.add_experimental_option("detach", True)
        # options.add_experimental_option('excludeSwitches', ['enable-logging'])
        # self.driver = webdriver.Chrome(options=options)

        self.driver.implicitly_wait(5)
        self.base_url = "https://www.google.com/"
        self.verificationErrors = []
        self.accept_next_alert = True

        # self.driver.find_element(By.XPATH,"").get_attribute(name="validationMessage")
        
        logger = logging.getLogger('urllib3.connectionpool')
        logger.setLevel(logging.INFO)

        logger = logging.getLogger('selenium.webdriver.remote.remote_connection')
        logger.setLevel(logging.WARNING)
        self.excel_file = openpyxl.load_workbook('testcase_input.xlsx').active
    # def testcase_1(self):
    #     filename=self.excel_file.cell(row=2, column=2).value
    #     expected = self.excel_file.cell(row=2, column=3).value
    #     testcaseRunner(self.driver, filename, expected)

    # def testcase_2(self):
    #     filename=self.excel_file.cell(row=3, column=2).value
    #     expected = self.excel_file.cell(row=3, column=3).value
    #     testcaseRunner(self.driver, filename, expected)

    # def testcase_3(self):
    #     filename=self.excel_file.cell(row=4, column=2).value
    #     expected = self.excel_file.cell(row=4, column=3).value
    #     testcaseRunner(self.driver, filename, expected)

    def testcase_4(self):
        filename=self.excel_file.cell(row=5, column=2).value
        expected = self.excel_file.cell(row=5, column=3).value
        testcaseRunner(self.driver, filename, expected)

    # def testcase_2(self):
    #     driver=self.driver
    #     driver.get("https://sandbox.moodledemo.net/")
    #     driver.find_element(By.XPATH, "/html/body/div[2]/nav/div[2]/div[5]/div/span/a").click()
    #     driver.find_element("id","username").click()
    #     driver.find_element("id","username").clear()
    #     driver.find_element("id","username").send_keys('student')

    #     driver.find_element("id","password").click()
    #     driver.find_element("id","password").clear()

    #     driver.find_element("id","password").send_keys('sandbox')
    #     driver.find_element("id","loginbtn").click()

    #     # driver.get("https://sandbox.moodledemo.net/user/edit.php?id=4&returnto=profile")
    #     driver.find_element("link text","My first course").click()
    #     driver.find_element("link text","test").click()

    #     driver.current_url

    #     driver.get("{}&action=editsubmission".format(driver.current_url))
    #     # dragdrop = driver.find_element_by_xpath('paste-the-full-xpath-here')
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,".fm-empty-container")))
    #         # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"/html/body/div[3]/div[4]/div/div[2]/div/section/div[2]/div/form/fieldset/div[2]/div/div[2]/fieldset/div[1]/div[4]/div[1]/div[2]")))
    #     dragdrop = driver.find_element(By.CSS_SELECTOR,".dndupload-target")

    #     cssSelector=".dndupload-message > div:nth-child(2)"
    #     # className='dndupload-arrow d-flex'
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR,cssSelector)))

    #     # ActionChains(driver).move_to_element(driver.find_element(By.CSS_SELECTOR,cssSelector)).perform()

    #     element= driver.find_element(By.CSS_SELECTOR,cssSelector).click()
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable(("link text","Upload a file")))
    #     driver.find_element("link text","Upload a file").click()
    #     # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.NAME,'repo_upload_file')))
    #     # driver.find_elements(By.NAME,'repo_upload_file')[0].click() --> not working
    #     # driver.find_elements(By.NAME,'repo_upload_file')[0].send_keys("img.png") 

    #     # driver.find_element("link text","Choose File").click()  #not working

    #     # driver.execute_script("document.getElementsByName('repo_upload_file')[0].click()") # working
    #     path = os.path.join(os.getcwd(),"test_data", "xlsfile.xls")
    #     # pyautogui.write(path, interval=0.25) 
    #     # # pyautogui.press('enter')
    #     # pyautogui.press('return')


    #     driver.find_element(By.NAME,'repo_upload_file' ).send_keys(path)
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]")))

    #     driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]").click()

    #     #Assert: 
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+'xlsfile.xls'+"')]")))
    #     driver.quit()

    # def testcase_3(self):
    #     driver=self.driver
    #     driver.get("https://sandbox.moodledemo.net/")
    #     driver.find_element(By.XPATH, "/html/body/div[2]/nav/div[2]/div[5]/div/span/a").click()
    #     driver.find_element("id","username").click()
    #     driver.find_element("id","username").clear()
    #     driver.find_element("id","username").send_keys('student')

    #     driver.find_element("id","password").click()
    #     driver.find_element("id","password").clear()

    #     driver.find_element("id","password").send_keys('sandbox')
    #     driver.find_element("id","loginbtn").click()

    #     # driver.get("https://sandbox.moodledemo.net/user/edit.php?id=4&returnto=profile")
    #     driver.find_element("link text","My first course").click()
    #     driver.find_element("link text","test").click()

    #     driver.current_url

    #     driver.get("{}&action=editsubmission".format(driver.current_url))
    #     # dragdrop = driver.find_element_by_xpath('paste-the-full-xpath-here')
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,".fm-empty-container")))
    #         # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"/html/body/div[3]/div[4]/div/div[2]/div/section/div[2]/div/form/fieldset/div[2]/div/div[2]/fieldset/div[1]/div[4]/div[1]/div[2]")))
    #     dragdrop = driver.find_element(By.CSS_SELECTOR,".dndupload-target")

    #     cssSelector=".dndupload-message > div:nth-child(2)"
    #     # className='dndupload-arrow d-flex'
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR,cssSelector)))

    #     # ActionChains(driver).move_to_element(driver.find_element(By.CSS_SELECTOR,cssSelector)).perform()

    #     element= driver.find_element(By.CSS_SELECTOR,cssSelector).click()
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable(("link text","Upload a file")))
    #     driver.find_element("link text","Upload a file").click()
    #     # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.NAME,'repo_upload_file')))
    #     # driver.find_elements(By.NAME,'repo_upload_file')[0].click() --> not working
    #     # driver.find_elements(By.NAME,'repo_upload_file')[0].send_keys("img.png") 

    #     # driver.find_element("link text","Choose File").click()  #not working

    #     # driver.execute_script("document.getElementsByName('repo_upload_file')[0].click()") # working
    #     path = os.path.join(os.getcwd(),"test_data", "xlsxfile.xlsx")
    #     # pyautogui.write(path, interval=0.25) 
    #     # # pyautogui.press('enter')
    #     # pyautogui.press('return')


    #     driver.find_element(By.NAME,'repo_upload_file' ).send_keys(path)
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]")))

    #     driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]").click()

    #     #Assert: 
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+'xlsxfile.xlsx'+"')]")))
    #     driver.quit()

 
    # def testcase_4(self):
    #     driver=self.driver
    #     driver.get("https://sandbox.moodledemo.net/")
    #     driver.find_element(By.XPATH, "/html/body/div[2]/nav/div[2]/div[5]/div/span/a").click()
    #     driver.find_element("id","username").click()
    #     driver.find_element("id","username").clear()
    #     driver.find_element("id","username").send_keys('student')

    #     driver.find_element("id","password").click()
    #     driver.find_element("id","password").clear()

    #     driver.find_element("id","password").send_keys('sandbox')
    #     driver.find_element("id","loginbtn").click()

    #     # driver.get("https://sandbox.moodledemo.net/user/edit.php?id=4&returnto=profile")
    #     driver.find_element("link text","My first course").click()
    #     driver.find_element("link text","test").click()

    #     driver.current_url

    #     driver.get("{}&action=editsubmission".format(driver.current_url))
    #     # dragdrop = driver.find_element_by_xpath('paste-the-full-xpath-here')
    #     self.driver.implicitly_wait(15)
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,".fm-empty-container")))
    #         # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"/html/body/div[3]/div[4]/div/div[2]/div/section/div[2]/div/form/fieldset/div[2]/div/div[2]/fieldset/div[1]/div[4]/div[1]/div[2]")))
    #     dragdrop = driver.find_element(By.CSS_SELECTOR,".dndupload-target")

    #     cssSelector=".dndupload-message > div:nth-child(2)"
    #     # className='dndupload-arrow d-flex'
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR,cssSelector)))

    #     # ActionChains(driver).move_to_element(driver.find_element(By.CSS_SELECTOR,cssSelector)).perform()

    #     element= driver.find_element(By.CSS_SELECTOR,cssSelector).click()
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable(("link text","Upload a file")))
    #     driver.find_element("link text","Upload a file").click()
    #     # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.NAME,'repo_upload_file')))
    #     # driver.find_elements(By.NAME,'repo_upload_file')[0].click() --> not working
    #     # driver.find_elements(By.NAME,'repo_upload_file')[0].send_keys("img.png") 

    #     # driver.find_element("link text","Choose File").click()  #not working

    #     # driver.execute_script("document.getElementsByName('repo_upload_file')[0].click()") # working
    #     path = os.path.join(os.getcwd(),"test_data", "img.png")
    #     # pyautogui.write(path, interval=0.25) 
    #     # # pyautogui.press('enter')
    #     # pyautogui.press('return')


    #     driver.find_element(By.NAME,'repo_upload_file' ).send_keys(path)
    #     driver.implicitly_wait(10) # seconds
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]"))).click()
    #     driver.implicitly_wait(10) # seconds
    #     # driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]").click()

    #     #Click `Upload profile`
    #     # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload profile'+"')]")))
    #     # driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload profile'+"')]").click()


    #     # Click upload button:
    #     # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID,"id_submitbutton")))
    #     # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+'img.png'+"')]")))
    #     # driver.find_element(By.ID,"id_submitbutton").click()

    #     #Assert: 
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+'Image (PNG) filetype cannot be accepted.'+"')]")))
    #     driver.quit()


    
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        # To know more about the difference between verify and assert,
        # visit https://www.seleniumhq.org/docs/06_test_design_considerations.jsp#validating-results
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    # unittest.main()
    runner = unittest.TextTestRunner(verbosity=2)
    # runner.run()
    unittest.main(testRunner=runner)
