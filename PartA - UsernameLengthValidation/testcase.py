# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions

from selenium.webdriver.chrome.service import Service
# import openpyxl module
import openpyxl
import logging
def runTestcase(driver, username, testcaseId ):
    driver.get("http://localhost/OnlinePizzaDelivery/")
    driver.find_element(By.XPATH,"//div[@id='navbarSupportedContent']/button[2]").click()
    driver.find_element(By.ID,"username").click()
    driver.find_element("id","username").send_keys(username)

    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//div[@id='signupModal']/div/div/div[2]/form/button")))
    driver.find_element(By.XPATH,"/html/body/div[2]/div/div/div[2]/form/button").click()
    element = driver.find_element("id","username")
    # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), 'Please lengthen this text')]")))
    validateStrr = element.get_attribute(name="validationMessage")

    # if expected not in validateStrr:
    if testcaseId==1:
        assert 'Please lengthen this text to 3 characters or more' in validateStrr
    elif testcaseId ==2:
        assert len(validateStrr)==0
    else:
        assert len(element.get_attribute('value')) ==11
    # print(strr)
    # driver.find_element("id","username").send_keys(validateStrr)
    driver.quit()


class AppDynamicsJob(unittest.TestCase):
    def setUp(self):
        # AppDynamics will automatically override this web driver
        # as documented in https://docs.appdynamics.com/display/PRO44/Write+Your+First+Script

        # options = FirefoxOptions()
        # options.set_preference('intl.accept_languages', 'en-US, en')
        # self.driver = webdriver.Firefox(options=options)


        options = ChromeOptions()
        options.add_experimental_option("detach", True)
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.driver = webdriver.Chrome(options=options)

        self.driver.implicitly_wait(5)
        self.base_url = "https://www.google.com/"
        self.verificationErrors = []
        self.accept_next_alert = True

        # self.driver.find_element(By.XPATH,"").get_attribute(name="validationMessage")
        
        logger = logging.getLogger('urllib3.connectionpool')
        logger.setLevel(logging.INFO)

        logger = logging.getLogger('selenium.webdriver.remote.remote_connection')
        logger.setLevel(logging.WARNING)
        # self.excel_file = openpyxl.load_workbook('testcase_input.xlsx').active
    def testcase_1(self):
        username='a'
        # expected = 'Please lengthen this text to 3 characters or more'
        runTestcase(self.driver, username, 1)

    def testcase_2(self):
        username='abcd'
        # expected = ''
        runTestcase(self.driver, username, 2)

    def testcase_3(self):
        username='abcdefgthikljmnopq'
        # expected = ''
        runTestcase(self.driver, username, 3)

    # def testcase_4(self):
    #     username=self.excel_file.cell(row=5, column=2).value
    #     password=self.excel_file.cell(row=5, column=3).value
    #     expected = self.excel_file.cell(row=5, column=4).value
    #     runTestcase(self.driver, username, password, expected)


    
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        # To know more about the difference between verify and assert,
        # visit https://www.seleniumhq.org/docs/06_test_design_considerations.jsp#validating-results
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    # unittest.main()
    runner = unittest.TextTestRunner(verbosity=2)
    # runner.run()
    unittest.main(testRunner=runner)
