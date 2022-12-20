# -*- coding: utf-8 -*-
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.action_chains import ActionChains

from selenium.webdriver.chrome.service import Service
# import openpyxl module
import openpyxl
import pyautogui

def runTestcase(driver, filename,  expected , tc_idx):
        driver.get("https://sandbox.moodledemo.net/")
        driver.find_element(By.XPATH, "/html/body/div[2]/nav/div[2]/div[5]/div/span/a").click()
        driver.find_element("id","username").click()
        driver.find_element("id","username").clear()
        driver.find_element("id","username").send_keys('student')

        driver.find_element("id","password").click()
        driver.find_element("id","password").clear()

        driver.find_element("id","password").send_keys('sandbox')
        driver.find_element("id","loginbtn").click()

        driver.get("https://sandbox.moodledemo.net/user/edit.php?id=4&returnto=profile")

        # xpath="//*[contains(text(), '{}')]".format(expected)
        # driver.find_element(By.XPATH,xpath)

        # try:
        #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,xpath)))
        # finally:
        #     # f.write("Assignment found\n")
        #     driver.quit()
        # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), 'Welcome back')]")))
        # expected='Welcome back'

        xpath='/html/body/div[3]/div[3]/div/div[2]/div/section/div/form/fieldset[2]/div[2]/div[2]/div[2]/fieldset/div[1]/div[4]/div[1]/div[2]/div/div'
        cssSelector=".dndupload-message > div:nth-child(2)"
        # className='dndupload-arrow d-flex'
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR,cssSelector)))

        # ActionChains(driver).move_to_element(driver.find_element(By.CSS_SELECTOR,cssSelector)).perform()

        element= driver.find_element(By.CSS_SELECTOR,cssSelector).click()
        try:
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.NAME,'repo_upload_file')))
        # driver.find_elements(By.NAME,'repo_upload_file')[0].click() --> not working
        # driver.find_elements(By.NAME,'repo_upload_file')[0].send_keys("img.png") 
        except:
            time.sleep(2)
            driver.execute_script("document.getElementsByName('repo_upload_file')[0].click()") # working

        if filename is not None:
            path = os.path.join(os.getcwd(),'test_data', filename)
            if not os.path.exists(path):
                raise Exception('Required data for this testcase not found!. Please check `current working directory.`')
            driver.find_element(By.NAME,'repo_upload_file' ).send_keys(path)

        # pyautogui.write(path, interval=0.25) 
        # # pyautogui.press('enter')
        # pyautogui.press('return')
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]")))

        driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]").click()

        #Click `Upload profile`
        # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload profile'+"')]")))
        # driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload profile'+"')]").click()
        if tc_idx==1: #Save
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID,"id_submitbutton")))
            driver.find_element(By.ID,"id_submitbutton").click()

        # if filename is not None:
        #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+filename+"')]")))
        #     driver.find_element(By.ID,"id_submitbutton").click()
        # else:
        #     driver.find_element(By.ID,"id_submitbutton").click()


        #Assert: 'Changes saved'
        time.sleep(5)
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+expected+"')]")))

        # driver.execute_script("arguments[0].scrollIntoView();", element)
        # driver.execute_script("document.querySelector('.dndupload-message > div:nth-child(2)').click()")
        # driver.implicitly_wait(19)
        # element.click()
        time.sleep(3)
        driver.quit()



class AppDynamicsJob(unittest.TestCase):
    def setUp(self):
        # AppDynamics will automatically override this web driver
        # as documented in https://docs.appdynamics.com/display/PRO44/Write+Your+First+Script

        # options = FirefoxOptions()
        # options.set_preference('intl.accept_languages', 'en-us')
        # self.driver = webdriver.Firefox(options=options)


        options = ChromeOptions()
        options.add_experimental_option("detach", True)
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.driver = webdriver.Chrome(options=options)

        self.driver.implicitly_wait(5)
        self.base_url = "https://www.google.com/"
        self.verificationErrors = []
        self.accept_next_alert = True

        self.excel_file = openpyxl.load_workbook('testcase_input.xlsx').active
    # def testcase_1(self):
    #     driver=self.driver
    #     driver.get("https://sandbox.moodledemo.net/")
    #     driver.find_element(By.XPATH, "/html/body/div[2]/nav/div[2]/div[5]/div/span/a").click()
    #     driver.find_element("id","username").click()
    #     driver.find_element("id","username").clear()
    #     driver.find_element("id","username").send_keys('student')

    #     driver.find_element("id","password").click()
    #     driver.find_element("id","password").clear()

    #     driver.find_element("id","password").send_keys('sandbox')
    #     driver.find_element("id","loginbtn").click()

    #     driver.get("https://sandbox.moodledemo.net/user/edit.php?id=4&returnto=profile")

    #     # xpath="//*[contains(text(), '{}')]".format(expected)
    #     # driver.find_element(By.XPATH,xpath)

    #     # try:
    #     #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,xpath)))
    #     # finally:
    #     #     # f.write("Assignment found\n")
    #     #     driver.quit()
    #     # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), 'Welcome back')]")))
    #     # expected='Welcome back'

    #     xpath='/html/body/div[3]/div[3]/div/div[2]/div/section/div/form/fieldset[2]/div[2]/div[2]/div[2]/fieldset/div[1]/div[4]/div[1]/div[2]/div/div'
    #     cssSelector=".dndupload-message > div:nth-child(2)"
    #     # className='dndupload-arrow d-flex'
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR,cssSelector)))

    #     # ActionChains(driver).move_to_element(driver.find_element(By.CSS_SELECTOR,cssSelector)).perform()

    #     element= driver.find_element(By.CSS_SELECTOR,cssSelector).click()
    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.NAME,'repo_upload_file')))
    #     # driver.find_elements(By.NAME,'repo_upload_file')[0].click() --> not working
    #     # driver.find_elements(By.NAME,'repo_upload_file')[0].send_keys("img.png") 


    #     # driver.execute_script("document.getElementsByName('repo_upload_file')[0].click()") # working
    #     path = os.path.join(os.getcwd(), "img.png")
    #     if not os.path.exists(path):
    #         raise Exception('Required data for this testcase not found!. Please check `current working directory.`')
    #     # pyautogui.write(path, interval=0.25) 
    #     # # pyautogui.press('enter')
    #     # pyautogui.press('return')


    #     driver.find_element(By.NAME,'repo_upload_file' ).send_keys(path)
    #     # driver.find_element(By.XPATH,'//body[1]/div[8]/div[3]/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]/div[1]/div[1]/form[1]/div[1]/div[1]/div[1]/input[1]').send_keys(path) 

    #     # driver.find_elements(By.CLASS_NAME,'fp-upload-btn btn-primary btn')[0].click() # Not working
    #     # document.getElementsByClassName('fp-upload-btn btn-primary btn')[0].click() #not test

    #     # driver.find_element(By.XPATH,'/html/body/div[8]/div[3]/div/div[2]/div/div/div[2]/div/div[2]/div/div/div/button').click()  not wrkoing

    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]")))

    #     driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload this file'+"')]").click()

    #     #Click `Upload profile`
    #     # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//*[contains(text(), '"+'Upload profile'+"')]")))
    #     # driver.find_element(By.XPATH,"//*[contains(text(), '"+'Upload profile'+"')]").click()

    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID,"id_submitbutton")))
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+'img.png'+"')]")))
    #     driver.find_element(By.ID,"id_submitbutton").click()

    #     #Assert: 'Changes saved'
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+'Changes saved'+"')]")))

    #     # driver.execute_script("arguments[0].scrollIntoView();", element)
        # driver.execute_script("document.querySelector('.dndupload-message > div:nth-child(2)').click()")
        # driver.implicitly_wait(19)
        # element.click()
        # driver.quit()



    # def testcase_1(self):
    #     filename=self.excel_file.cell(row=2, column=2).value
    #     expected = self.excel_file.cell(row=2, column=3).value
    #     runTestcase(self.driver, filename, expected,1)

    # def testcase_2(self):
    #     filename=self.excel_file.cell(row=3, column=2).value
    #     expected = self.excel_file.cell(row=3, column=3).value
    #     runTestcase(self.driver, filename, expected,2)

    
    def testcase_3(self):
        filename=self.excel_file.cell(row=4, column=2).value
        expected = self.excel_file.cell(row=4, column=3).value
        runTestcase(self.driver, filename, expected,2)

    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        # To know more about the difference between verify and assert,
        # visit https://www.seleniumhq.org/docs/06_test_design_considerations.jsp#validating-results
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    # unittest.main()
    runner = unittest.TextTestRunner(verbosity=2)
    # runner.run()
    unittest.main(testRunner=runner)
