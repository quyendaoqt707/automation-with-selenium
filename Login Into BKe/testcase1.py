# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions

from selenium.webdriver.chrome.service import Service
# import openpyxl module
import openpyxl

def runTestcase(driver, username, password, expected ):
    driver.get("https://sandbox.moodledemo.net/")
    driver.find_element(By.XPATH, "/html/body/div[2]/nav/div[2]/div[5]/div/span/a").click()
    driver.find_element("id","username").click()
    driver.find_element("id","username").clear()

    if username is not None:
        driver.find_element("id","username").send_keys(username)

    driver.find_element("id","password").click()
    driver.find_element("id","password").clear()

    if password is not None:
        driver.find_element("id","password").send_keys(password)
    driver.find_element("id","loginbtn").click()

    xpath="//*[contains(text(), '{}')]".format(expected)
    # driver.find_element(By.XPATH,xpath)

    # try:
    #     WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,xpath)))
    # finally:
    #     # f.write("Assignment found\n")
    #     driver.quit()
    # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), 'Welcome back')]")))
    # expected='Welcome back'
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(), '"+expected+"')]")))
    driver.quit()
    # driver.find_element(By.XPATH,"//*[contains(text(), 'Welcome back')]")

class AppDynamicsJob(unittest.TestCase):
    def setUp(self):
        # AppDynamics will automatically override this web driver
        # as documented in https://docs.appdynamics.com/display/PRO44/Write+Your+First+Script

        options = FirefoxOptions()
        options.set_preference('intl.accept_languages', 'en-us')
        self.driver = webdriver.Firefox(options=options)


        # options = ChromeOptions()
        # options.add_experimental_option("detach", True)
        # self.driver = webdriver.Chrome(options=options)

        self.driver.implicitly_wait(5)
        self.base_url = "https://www.google.com/"
        self.verificationErrors = []
        self.accept_next_alert = True

        self.excel_file = openpyxl.load_workbook('testcase_input.xlsx').active
    # def testcase_1(self):
    #     username=self.excel_file.cell(row=2, column=2).value
    #     password=self.excel_file.cell(row=2, column=3).value
    #     expected = self.excel_file.cell(row=2, column=4).value
    #     runTestcase(self.driver, username, password, expected)

    # def testcase_2(self):
    #     username=self.excel_file.cell(row=3, column=2).value
    #     password=self.excel_file.cell(row=3, column=3).value
    #     expected = self.excel_file.cell(row=3, column=4).value
    #     runTestcase(self.driver, username, password, expected)

    # def testcase_3(self):
    #     username=self.excel_file.cell(row=4, column=2).value
    #     password=self.excel_file.cell(row=4, column=3).value
    #     expected = self.excel_file.cell(row=4, column=4).value
    #     runTestcase(self.driver, username, password, expected)

    # def testcase_4(self):
    #     username=self.excel_file.cell(row=5, column=2).value
    #     password=self.excel_file.cell(row=5, column=3).value
    #     expected = self.excel_file.cell(row=5, column=4).value
    #     runTestcase(self.driver, username, password, expected)

    # def testcase_5(self):
    #     username=self.excel_file.cell(row=6, column=2).value
    #     password=self.excel_file.cell(row=6, column=3).value
    #     expected = self.excel_file.cell(row=6, column=4).value
    #     runTestcase(self.driver, username, password, expected)

    def testcase_6(self):
        username=self.excel_file.cell(row=7, column=2).value
        password=self.excel_file.cell(row=7, column=3).value
        expected = self.excel_file.cell(row=7, column=4).value
        runTestcase(self.driver, username, password, expected)


    def testcase_7(self):
        username=self.excel_file.cell(row=8, column=2).value
        password=self.excel_file.cell(row=8, column=3).value
        expected = self.excel_file.cell(row=8, column=4).value
        runTestcase(self.driver, username, password, expected)

    # def testcase_8(self):
    #     username=self.excel_file.cell(row=9, column=2).value
    #     password=self.excel_file.cell(row=9, column=3).value
    #     expected = self.excel_file.cell(row=9, column=4).value
    #     runTestcase(self.driver, username, password, expected)

    # def testcase_9(self):
    #     username=self.excel_file.cell(row=10, column=2).value
    #     password=self.excel_file.cell(row=10, column=3).value
    #     expected = self.excel_file.cell(row=10, column=4).value
    #     runTestcase(self.driver, username, password, expected)

    
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        # To know more about the difference between verify and assert,
        # visit https://www.seleniumhq.org/docs/06_test_design_considerations.jsp#validating-results
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    # unittest.main()
    runner = unittest.TextTestRunner(verbosity=2)
    # runner.run()
    unittest.main(testRunner=runner)
